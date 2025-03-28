import os, sys, time, re
import threading
import shutil

import nmap
import ipaddress
import socket

import datetime
import openpyxl

from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import messagebox

from grafico import Grafico
from velocidade import Velocidade
import interface
import variaveis

ultima_hora = None
horario = 22


def main():
    
    if shutil.which("nmap") is None:
        messagebox.showerror("Erro", "O NMAP não está instalado. Instale o Nmap antes de executar este programa.")
        sys.exit(1)
    messagebox.showwarning("Atenção!", "Esta é uma ferramenta utilizada para explorar a rede a fim de diagnosticar possíveis conflitos de IP em redes sem DHCP com vários dispositivos.\nEla coleta dados de scans feitos por NMAP com o tempo, e os organiza e salva em um arquivo chamado scan.xlsx.\nPara que a ferramenta funcione corretamente, tenha certeza de que ela está sendo executada em modo administrador para evitar potenciais erros.")
    variaveis.DEBUG = True
    root = tk.Tk()
    root.title("Scanner")
    #root.geometry("800x600")
    widgets = interface.Widgets(root)
    widgets.add_label("gateway", "Gateway Padrão:", 0, 0)
    widgets.add_entry("gateway", 1, 0, 4)
    widgets.add_label("prefix", "Prefixo:", 0, 2)
    widgets.add_entry("prefix", 1, 2)
    widgets.add_label("speed", "Velocidade(Mb):", 2, 2)
    widgets.add_entry("speed", 3, 2)
    widgets.add_label("options", "Opções de Scan:", 0, 3)
    widgets.add_entry("options", 1, 3)
    widgets.add_label("delay", "Delay(min):", 2, 3)
    widgets.add_entry("delay", 3, 3)
    widgets.add_checkbox("debug", "Debug", 0, 4, valor=variaveis.DEBUG)
    widgets.add_checkbox("log", "Log", 1, 4, valor=variaveis.LOG)
    widgets.add_button("start", "Iniciar", 0, 5, 4)
    widgets.add_text("text", "", 5, 0, 4)
    root.focus_force()
    threads = threading.Thread(target=root.mainloop())
    threads.start()
    widgets.entry_widget["gateway"].config({"text":"10.10.10.10"})
    while False:
        variaveis.RUNNING.wait()
        delay = widgets.entry_widget["delay"].get()
        alvo:str = widgets.entry_widget["gateway"].get()
        fim:int = widgets.entry_widget["prefix"].get()
        conf:str = widgets.entry_widget["options"].get()
        mapa:nmap.nmap.PortScanner



    local_arquivo:str = os.path.dirname(sys.executable)
    _ = re.split(r'/|\\', sys.executable)[-1].lower()
    if(_ == "python.exe"):
        local_arquivo = os.path.dirname(os.path.realpath(__file__))
    ARQUIVO_PLANILHA:str = local_arquivo + "/scan.xlsx"
    if(variaveis.LOG == True):
        print(sys.executable)
        print(ARQUIVO_PLANILHA)
        print(local_arquivo)
    dados:dict = inserir_dados()
    alvo:str = dados["alvo"]
    fim:int = dados["fim"]
    conf:str = dados["conf"]
    tempo:int = dados["tempo"]
    mapa:nmap.nmap.PortScanner
    validar_planilha(arquivo=ARQUIVO_PLANILHA, alvo=alvo, fim=fim)
    print(f"\n\n{"-"*50}\n\n")
    
    while True:
        now = datetime.datetime.now()
        print(f"{now.strftime("%H:%M:%S")}: Iniciando o scan em {alvo}")
        ips:dict = {}
        mapa = scan(alvo + "-" + str(fim), conf)
        #Adiciona os IPs e MACs escaneados à variável "ips", em seguida os ordena em ordem crescente
        for host in mapa.all_hosts():
            mac = mac_detalhe = "Nan"
            items = mapa[host]['vendor'].items()
            if(items):
                for _mac, _macdet in items:
                    mac = _mac
                    mac_detalhe = _macdet
            ips[host] = {"status":mapa[host].state(),"mac":mac,"detalhe":mac_detalhe}
        ips = dict(sorted(ips.items(), key=lambda item: int(ipaddress.ip_address(item[0]))))
        free_ip_lst:list = []
        prev = 1
        #Inicia o processo de identificação de endereços disponíveis
        for ip in ips:
            curr = int(ip.split('.')[-1])
            if(variaveis.LOG == True):
                print(f"{ip}\t{ips[ip]['status']}\tmac: {ips[ip]['mac'].ljust(20)}\tdetalhe: {ips[ip]['detalhe']}")
            if (curr>prev+1):
                free_ip_lst.extend(free_ip_handler(ip_alvo=alvo, ip_atual=curr, ip_anterior=prev))
            prev = curr
        #Caso o último IP em uso tenha o final menor que a variável <fim> (geralmente .255), realiza-se mais uma checagem adicionando todos os IPs entre o último em uso até <fim> (.255)
        if (prev<255):
            free_ip_lst.extend(free_ip_handler(ip_alvo=alvo, ip_atual=fim+1, ip_anterior=prev)) 
        if(variaveis.LOG == True):
            print(f"\n{"-"*50}\n\nLIVRES: \n")
            print(*free_ip_lst, sep='\n')
            print(len(ips))
        salvar_planilha(arquivo=ARQUIVO_PLANILHA, ips=ips, freeip=free_ip_lst, fim=fim, tempo=tempo)
        now = datetime.datetime.now()
        print(f"{now.strftime("%H:%M:%S")}: Scan em {alvo} realizado com sucesso. Próximo scan em {tempo} minuto(s).")
        if (variaveis.DEBUG == False):
            time.sleep(tempo*60)
        else:
            time.sleep(10)


def inserir_dados():
    alvo:str = ""
    fim:int = 255
    conf:str = "-sn" #sn Escaneia a rede mais rapidamente ao não buscar por portas. A remoção pode aumentar o número de resultados, mas irá aumentar drasticamente o tempo de execução da tarefa.
    tempo:int = 1

    while True:
        try:
            socket.socket
            alvo = input("Digite o ip do seu gateway padrão ou dispositivo que deseja verificar: Ex 192.168.1.1\n")
            ipaddress.IPv4Address(alvo)
            break
        except ValueError:
            print("Ip inválido.")
    print(f"\n{"-"*50}\n")
    while True:
        try:
            fim = int(input("Qual o final do último IP que deve ser escaneado? (pressione ENTER para o padrão 255)\n"))
            break
        except ValueError:
            print("IP final definido para 255")
            fim = 255
            break
    print(f"\n{"-"*50}\n")
    _inp = input("Digite o intervalo entre cada scan (em minutos): ex 10\n")
    if (_inp.isdigit()):
        tempo = int(_inp)
    else:
        tempo = 10
    print(f"\n{"-"*50}\n")
    _inp = input("Digite as opções de execução do nmap: Ex -sS -sn (ENTER para -sn)\n")
    print(f"\n{"-"*50}\n")
    if(_inp!=''):
        conf = _inp
    return {"alvo":alvo, "fim":fim, "tempo":tempo, "conf":conf}


def validar_planilha(arquivo, alvo, fim):
    #Seção onde ocorre manipulação inicial da planilha
    planilha = openpyxl.Workbook()
    try:
        if(os.path.exists(arquivo)):
            planilha = openpyxl.load_workbook(arquivo)
            _a_planilha = planilha["Status"]
            _ip = alvo.split(".")
            _ip[-1] = str(fim)
            _ipfinstr = ".".join(_ip)
            if(_a_planilha.cell(_a_planilha.max_row,2).value==_ipfinstr):
                print ("\nJá existe uma planilha com os valores necessários. Ela será usada.")
            else:
                raise Exception("A planilha não contém os valores necessários.")
        else:
            raise Exception("A planilha ainda não existe.")
    except Exception as e:
        if(os.path.exists(arquivo)):
            os.rename(arquivo, arquivo+".old")
        criar_planilha(local_planilha=arquivo, target=alvo, end=fim)
        print(f"Uma planilha onde serão salvos os dados foi criada em ''{arquivo}''")
    print("Lidando com a planilha...")
    try:
        planilha = openpyxl.load_workbook(arquivo)
        print("Planilha carregada com sucesso!")
    except Exception as e:
        input(f"A tentativa de carregar a planilha levantou o seguinte erro: {e}\nPressione ENTER para finalizar o programa")
        SystemExit()


def criar_planilha(local_planilha:str, target:str, end:int): #Formata a planilha e coloca o IP alvo na primeira célula

    wb = openpyxl.Workbook()
    statuswb = wb.active
    statuswb.title = "Status"
    wb.create_sheet("Grafico", 2)
    wb.create_sheet("Dados", 3)

    statuswb.column_dimensions['A'].width = 4
    statuswb.column_dimensions['B'].width = 20
    statuswb.column_dimensions['D'].width = 20
    statuswb.column_dimensions['E'].width = 150
    statuswb.cell(1,2, "IP")
    statuswb.cell(1,3, "STATUS")
    statuswb.cell(1,4, "TEMPO ONLINE")
    statuswb.cell(1,5, "MAC")
    for i, ip in enumerate(range(int(target.split('.')[-1]), end)): #Adiciona os IPs do escopo para a planilha e coloca o tempo de uso como 0
        i+=2 #+2 compensa as primeiras linhas que não possuem IPs
        statuswb.cell(i, 2, f"{".".join(target.split('.')[:-1])}.{str(ip+1)}")
        statuswb.cell(i, 4, "0")
    dadoswb = wb["Dados"]
    dadoswb.cell(1,1,"HORÁRIO")
    dadoswb.cell(1,2,"Online")
    dadoswb.cell(1,3,"Offline")
    dadoswb.cell(1,4,"Online")
    dadoswb.cell(1,5,"Offline")
    dadoswb.cell(1,6,"Conflituosos")
    dadoswb.cell(1,7,"Velocidade MB/s")
    dadoswb.cell(1,8,"Velocidade MB/s")
    for i in range(0,24):
        dadoswb.cell(i+2, 1, f"{i:02d}") #+2 para compensar o início em 0 do horário e index 
    wb.save(local_planilha)
    grafico = Grafico()
    grafico.setup(variaveis.DEBUG, variaveis.LOG, local_planilha)
    grafico.gerar_graficos()


def salvar_planilha(arquivo:str, ips:dict, freeip:list, fim:int, tempo:int):
    global ultima_hora
    global horario
   
    wb = openpyxl.load_workbook(arquivo)
    _conflito = 0
    if(variaveis.LOG == True):
        print(wb.sheetnames)
    wbstatus = wb["Status"]
    laranja = PatternFill(start_color="cf820e", end_color="cf820e", fill_type="solid")
    verde = PatternFill(start_color="68cf0e", end_color="68cf0e", fill_type="solid")
    vermelho = PatternFill(start_color="cf0e2b", end_color="cf0e2b", fill_type="solid")
    for linha in range(2,fim+1):#+2 compensa as primeiras linhas que não possuem IPs e o enumerate que inicia em 0
        wbstatus.cell(linha, 3, "")
        val = wbstatus.cell(row=linha,column=2).value
        if(val in freeip):
            wbstatus.cell(row=linha, column=1).fill = verde
        elif(val in ips):
            _status = ips[val]['status'].upper()
            wbstatus.cell(linha,3, _status)
            wbstatus.cell(linha,4, int(wbstatus.cell(row=linha, column=4).value) + tempo)
            wbstatus.cell(linha,4).alignment = openpyxl.styles.Alignment(horizontal='left')
            _celula = wbstatus.cell(row=linha, column=5)
            if(_celula.value != None and ips[val]['mac'] != "Nan"):
                if(ips[val]['mac'] not in _celula.value):
                    wbstatus.cell(linha, 5, wbstatus.cell(linha, 5).value + f" {ips[val]['mac']}")
            elif(ips[val]['mac'] != "Nan"):
                wbstatus.cell(linha, 5, f"{ips[val]['mac']}")
            if(_celula.value!=None):
                if (len(_celula.value.split(' '))>1):
                    wbstatus.cell(row=linha, column=1).fill = vermelho
                    _conflito += 1
                else:
                    wbstatus.cell(row=linha, column=1).fill = laranja
            else:
                wbstatus.cell(row=linha, column=1).fill = laranja

    _tempo = datetime.datetime.now().hour
    if(variaveis.DEBUG==True):
        _tempo = horario
        print(f"Horário(Debug): {_tempo}")
    if (_tempo != ultima_hora):
        wbdados = wb["Dados"]
        if(_tempo == 0 and ultima_hora != None):
            wbdados.move_range("B2:C25", cols=2)
            wbdados.move_range("G2:G25", cols=1)
            for linha in range(2,26):
                wbdados[f"F{linha}"].value = None
                wbdados[f"G{linha}"].value = None
        ultima_hora = _tempo
        horario = horario+1 if horario < 23 else 0 #debug, don't care about it
        wbdados.cell(_tempo+2, 2, len(ips)) #Somamos 2 ao tempo para compensar a primeira linha de informações e o início das horas em 0
        wbdados.cell(_tempo+2, 3, len(freeip))
        wbdados.cell(_tempo+2, 6, _conflito)
        print(f"Testando a velocidade da internet...")
        _velocidade = Velocidade().teste(variaveis.LOG)
        _velocidade = (_velocidade/1000000)/8 #/8 para mB
        _velocidade = round(_velocidade, 3)
        wbdados.cell(_tempo+2, 7, _velocidade)
    wb.save(arquivo)

def scan(target, conf): #Responsável por executar o NMAP no IP fornecido
    resultado = nmap.PortScanner()
    try:
        resultado.scan(target, arguments=conf)
    except Exception as e:
        input(f"Ocorreu o erro '{e}' ao tentar escanear a rede.\nPressione ENTER para finalizar o programa e tente novamente com outras configurações.\n")
        SystemExit()
    return resultado


def free_ip_handler(ip_alvo:str, ip_atual:int, ip_anterior:int): #Detecta quais IPs estão livres e os retorna
    _ips = []
    for i in range(ip_anterior+1, ip_atual):
        base = ip_alvo.split(".")
        _ips.append(f"{".".join(base[:-1])}.{str(i)}")
    return _ips


if(__name__ == "__main__"):
    main()