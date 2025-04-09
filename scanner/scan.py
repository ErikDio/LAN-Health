import datetime, time, os, re, sys
import threading
import tkinter as tk

import ipaddress
import nmap

import openpyxl
from openpyxl.styles import PatternFill

from grafico import Grafico
from velocidade import Velocidade
import interface
import variaveis
import log

class Scan():
    def __init__(self, gateway:str, prefix:int, delay:int, speed:int, config:str):
        self.delay:int = delay
        self.gateway:str = gateway
        self.prefix:int = prefix
        self.speed:int = speed
        self.config:str = config
        self.mapa:nmap.nmap.PortScanner
        local_arquivo:str = os.path.dirname(sys.executable)
        _ = re.split(r'/|\\', sys.executable)[-1].lower()
        if(_ == "python.exe"):
            local_arquivo = os.path.dirname(os.path.realpath(__file__))
        self.arquivo_planilha:str = local_arquivo + f"/scan-{self.alvo.replace('.', '-')}.xlsx"
        self.planilha:openpyxl.Workbook = self.validar_planilha()
        log.box_text_log(f"{self.gateway}/{self.prefix} {self.config} - {self.delay} minutos - {self.speed} MB/s")
        self.lista_ips = self.gerar_ips()
        self.lista_ips.remove(self.gateway)
        log.box_text_log(f"IPs gerados: {', '.join(self.lista_ips)}")

    def gerar_ips(self):
        return [str(ip) for ip in ipaddress.IPv4Network(f"{self.gateway}/{self.prefix}", strict=False)]
    
    def run(self):
        while True:
            if (variaveis.FINISHING.is_set() == True):
                variaveis.ABORTED.set()
                return
            now = datetime.datetime.now()
            print(f"{now.strftime("%H:%M:%S")}: Iniciando o scan em {self.gateway}")
            ips:dict = {}
            self.mapa = self.nmap_scan(f"{self.gateway}/{self.prefix}", self.conf)
            #Adiciona os IPs e MACs escaneados à variável "ips", em seguida os ordena em ordem crescente
            for host in self.mapa.all_hosts():
                mac = mac_detalhe = "Nan"
                items = self.mapa[host]['vendor'].items()
                if(items):
                    for _mac, _macdet in items:
                        mac = _mac
                        mac_detalhe = _macdet
                ips[host] = {"status":self.mapa[host].state(),"mac":mac,"detalhe":mac_detalhe}
            ips = dict(sorted(ips.items(), key=lambda item: int(ipaddress.ip_address(item[0]))))
            free_ip_lst:list = []
            prev = 1
            for ip in ips:
                curr = int(ip.split('.')[-1])
                if(variaveis.LOG == True):
                    print(f"{ip}\t{ips[ip]['status']}\tmac: {ips[ip]['mac'].ljust(20)}\tdetalhe: {ips[ip]['detalhe']}")
                if (curr>prev+1):
                    free_ip_lst.extend(free_ip_handler(ip_alvo=alvo, ip_atual=curr, ip_anterior=prev))
                prev = curr
            #Caso o último IP em uso tenha o final menor que a variável <fim> (geralmente .255), realiza-se mais uma checagem adicionando todos os IPs entre o último em uso até <fim> (.255)
            if (prev<255):
                free_ip_lst.extend(free_ip_handler(ip_alvo=alvo, ip_atual=fim+1, ip_anterior=prev))
            if(variaveis.LOG == True):
                print(f"\n{"-"*50}\n\nLIVRES: \n")
                print(*free_ip_lst, sep='\n')
                print(len(ips))
            salvar_planilha(arquivo=ARQUIVO_PLANILHA, ips=ips, freeip=free_ip_lst, fim=fim, tempo=tempo)
            now = datetime.datetime.now()
            print(f"{now.strftime("%H:%M:%S")}: Scan em {alvo} realizado com sucesso. Próximo scan em {tempo} minuto(s).")
            threading.Event.wait(variaveis.FINISHING, self.delay*60)

    def validar_planilha(self):
    #Seção onde ocorre manipulação inicial da planilha
        planilha = openpyxl.Workbook()
        if(os.path.exists(self.arquivo_planilha)):
            pass
        else:
            log.box_text_log("A planilha ainda não existe.", True)
            self.criar_planilha()
            log.box_text(f"Uma planilha onde serão salvos os dados foi criada em ''{self.arquivo_planilha}''", True)
        log.box_text_log("Lidando com a planilha...")
        try:
            planilha = openpyxl.load_workbook(self.arquivo)
            log.box_text_log(f"Planilha scan-{self.alvo.replace('.', '-')}.xlsx carregada com sucesso!", True)
            return planilha
        except Exception as e:
            input(f"A tentativa de carregar a planilha levantou o seguinte erro: {e}\nA planilha pode estar corrompida ou não existe. Por favor, verifique o arquivo e tente novamente.")
            SystemExit()


    def criar_planilha(self, local_planilha:str, target:str, end:int): #Formata a planilha e coloca o IP alvo na primeira célula
        wb = openpyxl.Workbook()
        statuswb = wb.active
        statuswb.title = "Status"
        wb.create_sheet("Grafico", 2)
        wb.create_sheet("Dados", 3)

        statuswb.column_dimensions['A'].width = 4
        statuswb.column_dimensions['B'].width = 20
        statuswb.column_dimensions['D'].width = 20
        statuswb.column_dimensions['E'].width = 150
        statuswb.cell(1,2, "IP")
        statuswb.cell(1,3, "STATUS")
        statuswb.cell(1,4, "TEMPO ONLINE")
        statuswb.cell(1,5, "MAC")
        for i, ip in enumerate(self.lista_ips): #Adiciona os IPs do escopo para a planilha e coloca o tempo de uso como 0
            i+=2 #+2 compensa as primeiras linhas que não possuem IPs
            statuswb.cell(i, 2, f"{ip}")
            statuswb.cell(i, 4, "0")
        dadoswb = wb["Dados"]
        dadoswb.cell(1,1,"HORÁRIO")
        dadoswb.cell(1,2,"Online")
        dadoswb.cell(1,3,"Offline")
        dadoswb.cell(1,4,"Online")
        dadoswb.cell(1,5,"Offline")
        dadoswb.cell(1,6,"Conflituosos")
        dadoswb.cell(1,7,"Velocidade MB/s")
        dadoswb.cell(1,8,"Velocidade MB/s")
        for i in range(0,24):
            dadoswb.cell(i+2, 1, f"{i:02d}") #+2 para compensar o início em 0 do horário e index 
        wb.save(self.arquivo_planilha)
        grafico = Grafico()
        grafico.setup(self.arquivo_planilha)
        grafico.gerar_graficos()


    def salvar_planilha(self, arquivo:str, ips:dict, freeip:list, fim:int, tempo:int):
        global ultima_hora
        global horario
    
        wb = openpyxl.load_workbook(arquivo)
        _conflito = 0
        if(variaveis.LOG == True):
            print(wb.sheetnames)
        wbstatus = wb["Status"]
        laranja = PatternFill(start_color="cf820e", end_color="cf820e", fill_type="solid")
        verde = PatternFill(start_color="68cf0e", end_color="68cf0e", fill_type="solid")
        vermelho = PatternFill(start_color="cf0e2b", end_color="cf0e2b", fill_type="solid")
        for linha in range(2,fim+1):#+2 compensa as primeiras linhas que não possuem IPs e o enumerate que inicia em 0
            wbstatus.cell(linha, 3, "")
            val = wbstatus.cell(row=linha,column=2).value
            if(val in freeip):
                wbstatus.cell(row=linha, column=1).fill = verde
            elif(val in ips):
                _status = ips[val]['status'].upper()
                wbstatus.cell(linha,3, _status)
                wbstatus.cell(linha,4, int(wbstatus.cell(row=linha, column=4).value) + tempo)
                wbstatus.cell(linha,4).alignment = openpyxl.styles.Alignment(horizontal='left')
                _celula = wbstatus.cell(row=linha, column=5)
                if(_celula.value != None and ips[val]['mac'] != "Nan"):
                    if(ips[val]['mac'] not in _celula.value):
                        wbstatus.cell(linha, 5, wbstatus.cell(linha, 5).value + f" {ips[val]['mac']}")
                elif(ips[val]['mac'] != "Nan"):
                    wbstatus.cell(linha, 5, f"{ips[val]['mac']}")
                if(_celula.value!=None):
                    if (len(_celula.value.split(' '))>1):
                        wbstatus.cell(row=linha, column=1).fill = vermelho
                        _conflito += 1
                    else:
                        wbstatus.cell(row=linha, column=1).fill = laranja
                else:
                    wbstatus.cell(row=linha, column=1).fill = laranja

        _tempo = datetime.datetime.now().hour
        if(variaveis.DEBUG==True):
            _tempo = horario
            print(f"Horário(Debug): {_tempo}")
        if (_tempo != ultima_hora):
            wbdados = wb["Dados"]
            if(_tempo == 0 and ultima_hora != None):
                wbdados.move_range("B2:C25", cols=2)
                wbdados.move_range("G2:G25", cols=1)
                for linha in range(2,26):
                    wbdados[f"F{linha}"].value = None
                    wbdados[f"G{linha}"].value = None
            ultima_hora = _tempo
            horario = horario+1 if horario < 23 else 0 #debug, don't care about it
            wbdados.cell(_tempo+2, 2, len(ips)) #Somamos 2 ao tempo para compensar a primeira linha de informações e o início das horas em 0
            wbdados.cell(_tempo+2, 3, len(freeip))
            wbdados.cell(_tempo+2, 6, _conflito)
            log.box_text_log(f"Testando a velocidade da internet...")
            _velocidade = Velocidade().teste(variaveis.LOG)
            _velocidade = (_velocidade/1000000)/8 #/8 para mB
            _velocidade = round(_velocidade, 3)
            wbdados.cell(_tempo+2, 7, _velocidade)
        wb.save(arquivo)
    def nmap_scan(target, conf): #Responsável por executar o NMAP no IP fornecido
        resultado = nmap.PortScanner()
        try:
            resultado.scan(target, arguments=conf)
        except Exception as e:
            log.crash(f"Ocorreu o erro '{e}' ao tentar escanear a rede.\nCaso o erro persista, verifique se o nmap está corretamente instalado e se o programa tem privilégios suficiente.")
            
        return resultado