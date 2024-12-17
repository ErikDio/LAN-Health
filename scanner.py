import nmap
import ipaddress
import openpyxl
import os
import sys
import re
import time
import datetime
from openpyxl.styles import PatternFill

DEBUG:bool = True

def main():
    global DEBUG
    try:
        if(sys.argv[1].lower() == "debug"):
            DEBUG = True
        else:
            DEBUG = False
    except IndexError:
        DEBUG = False
    print("\nEsta é uma ferramenta utilizada para explorar a rede a fim de diagnosticar possíveis conflitos de IP em redes sem DHCP com vários dispositivos.\nEla coleta dados de scans feitos por NMAP com o tempo, e os organiza e salva em um arquivo chamado result.xlsx.\nPara que a ferramenta funcione corretamente, tenha certeza de que ela está sendo executada em modo administrador para evitar potenciais erros.")
    local_arquivo:str = os.path.dirname(sys.executable)
    _ = re.split(r'/|\\', sys.executable)[-1].lower()
    if(_ == "python.exe"):
        local_arquivo = os.path.dirname(os.path.realpath(__file__))
    arquivo_planilha:str = local_arquivo + "/plan.xlsx"
    if(DEBUG == True):
        print(sys.executable)
        print(arquivo_planilha)
        print(local_arquivo)
    print(f"\n{"-"*50}\n")
    alvo:str = ""
    fim:int = 255
    mapa:nmap.nmap.PortScanner
    conf:str = "-sn" #sn Escaneia a rede mais rapidamente ao não buscar por portas. A remoção pode aumentar o número de resultados, mas irá aumentar drasticamente o tempo de execução da tarefa.
    planilha = openpyxl.Workbook()
    tempo = 1
    while True:
        try:
            alvo = input("Digite o ip do seu gateway padrão ou dispositivo que deseja verificar: Ex 192.168.1.1\n")
            ipaddress.IPv4Address(alvo)
            break
        except ValueError:
            print("Ip inválido.")
    print(f"\n{"-"*50}\n")
    while True:
        try:
            fim = int(input("Qual o final do último IP que deve ser escaneado? (pressione ENTER para o padrão 255)\n"))
            break
        except ValueError:
            print("IP final definido para 255")
            fim = 255
            break
    print(f"\n{"-"*50}\n")
    while True:
        _inp = input("Digite o intervalo entre cada scan (em minutos): ex 10\n")
        if (_inp.isdigit()):
            tempo = int(_inp)
            break
    print(f"\n{"-"*50}\n")
    _inp = input("Digite as opções de execução do nmap: Ex -sS -sn (ENTER para -sn)\n")
    print(f"\n{"-"*50}\n")
    if(_inp!=''):
        conf = _inp
    #Seção onde ocorre manipulação inicial da planilha
    try:
        if(os.path.exists(arquivo_planilha)):
            planilha = openpyxl.load_workbook(arquivo_planilha)
            _a_planilha = planilha.active
            if(_a_planilha.cell(1,1).value==alvo):
                print ("\nPlanilha contém o valor correto de cabeçalho.")
            else:
                raise Exception("Arquivo não contém o mesmo IP no cabeçalho.")
        else:
            raise Exception("Arquivo ainda não existe.")
    except Exception as e:
        if(os.path.exists(arquivo_planilha)):
            os.rename(arquivo_planilha, arquivo_planilha+".old")
        criar_planilha(local_planilha=arquivo_planilha, target=alvo, end=fim)
        print(f"Uma planilha onde serão salvos os dados foi criada em ''{arquivo_planilha}''")
    print("Lidando com a planilha...")
    try:
        planilha = openpyxl.load_workbook(arquivo_planilha)
        print("Planilha carregada com sucesso!")
    except Exception as e:
        input(f"A tentativa de carregar a planilha levantou o seguinte erro: {e}\nPressione ENTER para finalizar o programa")
        return
    
    print(f"\n\n{"-"*50}\n\n")
    while True:
        now = datetime.datetime.now()
        print(f"{now.strftime("%H:%M:%S")}: Iniciando o scan em {alvo}")
        ips:dict = {}
        mapa = scan(alvo + "-" + str(fim), conf)
        #Adiciona os IPs e MACs escaneados à variável "ips", em seguida os ordena em ordem crescente
        for host in mapa.all_hosts():
            mac = mac_detalhe = "Nan"
            items = mapa[host]['vendor'].items()
            if(items):
                for _mac, _macdet in items:
                    mac = _mac
                    mac_detalhe = _macdet
            ips[host] = {"status":mapa[host].state(),"mac":mac,"detalhe":mac_detalhe}
        ips = dict(sorted(ips.items(), key=lambda item: int(ipaddress.ip_address(item[0]))))
        free_ip_lst:list = []
        prev = 1
        #Inicia o processo de identificação de endereços disponíveis
        for ip in ips:
            curr = int(ip.split('.')[-1])
            if(DEBUG == True):
                print(f"{ip}\t{ips[ip]['status']}\tmac: {ips[ip]['mac'].ljust(20)}\tdetalhe: {ips[ip]['detalhe']}")
            if (curr>prev+1):
                free_ip_lst.extend(free_ip_handler(ip_alvo=alvo, ip_atual=curr, ip_anterior=prev))
            prev = curr
        #Caso o último IP em uso tenha o final menor que a variável <fim> (geralmente .255), realiza-se mais uma checagem adicionando todos os IPs entre o último em uso até <fim> (.255)
        if (prev<255):
            free_ip_lst.extend(free_ip_handler(ip_alvo=alvo, ip_atual=fim+1, ip_anterior=prev)) 
        if(DEBUG == True):
            print(f"\n{"-"*50}\n\nLIVRES: \n")
            print(*free_ip_lst, sep='\n')
            print(len(ips))
        salvar_planilha(arquivo=arquivo_planilha, ips=ips, freeip=free_ip_lst, fim=fim, tempo=tempo)
        now = datetime.datetime.now()
        print(f"{now.strftime("%H:%M:%S")}: Scan em {alvo} realizado com sucesso. Próximo scan em {tempo} minuto(s).")
        time.sleep(tempo*60)

def criar_planilha(local_planilha:str, target:str, end:int): #Formata a planilha e coloca o IP alvo na primeira célula
    wb = openpyxl.Workbook()
    awb = wb.active
    global DEBUG
    if(DEBUG == True):
        print(awb['F1'].value)
    awb.merge_cells("A1:E1")
    awb.column_dimensions['A'].width = 4
    awb.column_dimensions['B'].width = 20
    awb.column_dimensions['D'].width = 20
    awb.column_dimensions['E'].width = 200
    awb.cell(1,1, target)
    awb.cell(2,2, "IP")
    awb.cell(2,3, "STATUS")
    awb.cell(2,4, "TEMPO ONLINE")
    awb.cell(2,5, "MAC")
    for i, ip in enumerate(range(int(target.split('.')[-1]), end)): #Adiciona os IPs do escopo para a planilha e coloca o tempo de uso como 0
        awb.cell(i+3, 2, f"{".".join(target.split('.')[:-1])}.{str(ip+1)}")
        awb.cell(i+3, 4, "0")
    wb.save(local_planilha)

def salvar_planilha(arquivo:str, ips:dict, freeip:list, fim:int, tempo:int):
    wb = openpyxl.load_workbook(arquivo)
    awb = wb.active
    laranja = PatternFill(start_color="cf820e", end_color="cf820e", fill_type="solid")
    verde = PatternFill(start_color="68cf0e", end_color="68cf0e", fill_type="solid")
    vermelho = PatternFill(start_color="cf0e2b", end_color="cf0e2b", fill_type="solid")
    for linha in range(3,fim+2):#+2 compensa as primeiras duas linhas que não possuem IPs
        awb.cell(linha, 3, "")
        val = awb.cell(row=linha,column=2).value
        if(val in freeip):
            awb.cell(row=linha, column=1).fill = verde
        elif(val in ips):
            _status = ips[val]['status'].upper()
            awb.cell(linha,3, _status)
            awb.cell(linha,4, int(awb.cell(row=linha, column=4).value) + tempo)
            awb.cell(linha,4).alignment = openpyxl.styles.Alignment(horizontal='left')
            _celula = awb.cell(row=linha, column=5)
            if(_celula.value != None and ips[val]['mac'] != "Nan"):
                if(ips[val]['mac'] not in _celula.value):
                    awb.cell(linha, 5, awb.cell(linha, 5).value + f" {ips[val]['mac']}")
            elif(ips[val]['mac'] != "Nan"):
                awb.cell(linha, 5, f"{ips[val]['mac']}")
            if(_celula.value!=None):
                if (len(_celula.value.split(' '))>1):
                    awb.cell(row=linha, column=1).fill = vermelho
                else:
                    awb.cell(row=linha, column=1).fill = laranja
            else:
                awb.cell(row=linha, column=1).fill = laranja
    wb.save(arquivo)


def scan(target, conf): #Responsável por executar o NMAP no IP fornecido
    resultado = nmap.PortScanner()
    try:
        resultado.scan(target, arguments=conf)
    except Exception as e:
        input(f"Ocorreu o erro '{e}' ao tentar escanear a rede.\nPressione ENTER para finalizar o programa e tente novamente com outras configurações.\n")
        SystemExit()
    return resultado

def free_ip_handler(ip_alvo:str, ip_atual:int, ip_anterior:int): #Detecta quais IPs estão livres e os retorna
    _ips = []
    for i in range(ip_anterior+1, ip_atual):
        base = ip_alvo.split(".")
        _ips.append(f"{".".join(base[:-1])}.{str(i)}")
    return _ips


if(__name__ == "__main__"):
    main()
